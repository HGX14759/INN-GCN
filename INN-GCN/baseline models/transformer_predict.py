import os
import numpy as np
import pandas as pd
import joblib
import torch
import torch.nn as nn
from rdkit import Chem, RDLogger
from rdkit.Chem import AllChem
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

BEST_PARAMS = {
    'd_model': 128,
    'nhead': 4,
    'num_layers': 1,
    'dropout': 0.12183907996330846,
    'lr': 0.00015605916948313548
}

DATASET_PATH = r"D:\project\predict\sjyc-csj.xlsx"
MODEL_PATH = r"D:\project\contrast model\cocrystal_prediction_Transformer\ensemble_transformer_final.pkl"
OUTPUT_PATH = r"D:\project\predict\transform_prediction_result.xlsx"

ECFP_RADIUS = 2
ECFP_NBITS = 1024

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

BASE_SEED = 42
np.random.seed(BASE_SEED)
torch.manual_seed(BASE_SEED)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(BASE_SEED)

RDLogger.DisableLog('rdApp.*')

class FeatureTransformer(nn.Module):
    def __init__(self, input_dim, d_model=128, nhead=4, num_layers=2, dropout=0.1):
        super().__init__()
        self.proj = nn.Linear(input_dim, d_model)
        self.pos_emb = nn.Parameter(torch.randn(1, 1, d_model))

        encoder_layer = nn.TransformerEncoderLayer(
            d_model=d_model, nhead=nhead, dim_feedforward=d_model * 2,
            dropout=dropout, batch_first=True, activation='gelu'
        )
        self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=num_layers)

        self.norm = nn.LayerNorm(d_model)
        self.head = nn.Sequential(
            nn.Linear(d_model, d_model // 2),
            nn.GELU(),
            nn.Dropout(dropout),
            nn.Linear(d_model // 2, 1)
        )

    def forward(self, x):
        x = self.proj(x).unsqueeze(1)
        x = x + self.pos_emb
        x = self.transformer(x)
        x = self.norm(x).squeeze(1)
        return torch.sigmoid(self.head(x)).squeeze(-1)


class EnsembleModel:
    def __init__(self, models, scalers):
        self.models = models
        self.scalers = scalers
        for m in self.models:
            m.eval()

    @torch.no_grad()
    def predict_proba(self, X):
        all_probs = []
        for model, sc in zip(self.models, self.scalers):
            x = sc.transform(X)
            x = torch.tensor(x, dtype=torch.float32).to(DEVICE)
            prob = model(x).cpu().numpy()
            all_probs.append(prob)
        return np.mean(all_probs, axis=0)

def smiles_to_ecfp4(smiles):
    """将SMILES转换为ECFP4指纹，无效SMILES返回全0向量"""
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return np.zeros(ECFP_NBITS, dtype=np.float32)
    fp = AllChem.GetMorganFingerprintAsBitVect(mol, radius=ECFP_RADIUS, nBits=ECFP_NBITS)
    return np.array(fp, dtype=np.float32)


def calculate_complementary_features(api_desc, ccf_desc, api_hba_hbd_homo_lumo, ccf_hba_hbd_homo_lumo):
    """计算分子配对互补特征"""
    api_hba, api_hbd, api_homo, api_lumo = api_hba_hbd_homo_lumo
    ccf_hba, ccf_hbd, ccf_homo, ccf_lumo = ccf_hba_hbd_homo_lumo
    api_rbn, api_s, api_s_l, api_s_m, api_m_l, api_fr_no, api_fr_aromaticAtom, \
        api_xlogp3, api_tpsa, api_acd_logp, api_mv, api_polarizability, api_dipole = api_desc
    ccf_rbn, ccf_s, ccf_s_l, ccf_s_m, ccf_m_l, ccf_fr_no, ccf_fr_aromaticAtom, \
        ccf_xlogp3, ccf_tpsa, ccf_acd_logp, ccf_mv, ccf_polarizability, ccf_dipole = ccf_desc

    pair_feat = [
        api_hba * ccf_hbd + api_hbd * ccf_hba,
        api_homo - ccf_lumo,
        ccf_homo - api_lumo,
        api_polarizability * ccf_polarizability,
        abs(api_rbn - ccf_rbn), abs(api_s - ccf_s), abs(api_s_l - ccf_s_l),
        abs(api_s_m - ccf_s_m), abs(api_m_l - ccf_m_l), abs(api_fr_no - ccf_fr_no),
        abs(api_fr_aromaticAtom - ccf_fr_aromaticAtom), abs(api_xlogp3 - ccf_xlogp3),
        abs(api_tpsa - ccf_tpsa), abs(api_acd_logp - ccf_acd_logp),
        abs(api_mv - ccf_mv), abs(api_dipole - ccf_dipole)
    ]
    return np.array(pair_feat, dtype=np.float32)


def generate_features_from_dataset(df):
    """从数据集生成完整的特征矩阵"""
    X_list = []
    for idx, row in df.iterrows():
        api_fp = smiles_to_ecfp4(row['SMILES1'])
        ccf_fp = smiles_to_ecfp4(row['SMILES2'])

        api_17 = [row[c] for c in [
            'API_RBN', 'API_S', 'API_S_L', 'API_S_M', 'API_M_L', 'API_Fr_NO', 'API_Fr_aromaticAtom',
            'API_XLogP3', 'API_Topological Polar Surface Area', 'API_ACD/LogP', 'API_MV',
            'API_Polarizability', 'API_Dipole Moment', 'API_HBA', 'API_HBD', 'API_homo', 'API_lumo'
        ]]
        ccf_17 = [row[c] for c in [
            'CCF_RBN', 'CCF_S', 'CCF_S_L', 'CCF_S_M', 'CCF_M_L', 'CCF_Fr_NO', 'CCF_Fr_aromaticAtom',
            'CCF_XLogP3', 'CCF_Topological Polar Surface Area', 'CCF_ACD/LogP', 'CCF_MV',
            'CCF_Polarizability', 'CCF_Dipole Moment', 'CCF_HBA', 'CCF_HBD', 'CCF_homo', 'CCF_lumo'
        ]]

        api13, api4 = api_17[:13], api_17[13:]
        ccf13, ccf4 = ccf_17[:13], ccf_17[13:]

        pair_feat = calculate_complementary_features(api13, ccf13, api4, ccf4)

        full_feat = np.concatenate([api_fp, ccf_fp, pair_feat])
        X_list.append(full_feat)

    return np.array(X_list, dtype=np.float32)

def main():
    print("="*60)
    print("晶型预测模型 - 全新数据集预测脚本")
    print("="*60)
    print(f"固定最优超参数: {BEST_PARAMS}")
    print(f"数据集路径: {DATASET_PATH}")
    print(f"模型路径: {MODEL_PATH}")
    print("="*60)

    print("\n1. 加载数据集...")
    try:
        df = pd.read_excel(DATASET_PATH, sheet_name='Sheet1')
        print(f"✅ 数据集加载成功，共 {len(df)} 条数据")
    except Exception as e:
        print(f"❌ 数据集加载失败: {e}")
        return

    print("\n2. 生成分子特征...")
    try:
        X = generate_features_from_dataset(df)
        print(f"✅ 特征生成完成，特征维度: {X.shape}")
    except Exception as e:
        print(f"❌ 特征生成失败: {e}")
        return

    print("\n3. 加载训练好的集成模型...")
    try:
        ensemble_model = joblib.load(MODEL_PATH)
        print(f"✅ 模型加载成功，共 {len(ensemble_model.models)} 个基模型")
    except Exception as e:
        print(f"❌ 模型加载失败: {e}")
        print("请确保模型文件路径正确，且已完成模型训练")
        return

    print("\n4. 执行预测...")
    try:
        y_pred_proba = ensemble_model.predict_proba(X)
        y_pred = (y_pred_proba > 0.5).astype(int)
        print("✅ 预测完成")
    except Exception as e:
        print(f"❌ 预测失败: {e}")
        return

    print("\n5. 整理预测结果...")
    result_df = df.copy()
    result_df['预测正类概率'] = y_pred_proba.round(4)
    result_df['预测标签'] = y_pred
    result_df['预测标签说明'] = result_df['预测标签'].map({0: '无法形成共晶', 1: '可以形成共晶'})

    print("\n6. 保存预测结果...")
    try:
        with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='预测结果', index=False)

        wb = load_workbook(OUTPUT_PATH)
        ws = wb['预测结果']

        header_fill = PatternFill('solid', fgColor='0070C0')
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        thin_gray = Side(style='thin', color='D9D9D9')
        full_border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

        zebra_blue = 'EBF1F8'
        white = 'FFFFFF'

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = full_border
        ws.row_dimensions[1].height = 30

        for row in range(2, ws.max_row + 1):
            bg = zebra_blue if (row - 2) % 2 == 0 else white
            fill = PatternFill('solid', fgColor=bg)

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = full_border

                col_name = ws.cell(row=1, column=col).value
                if col_name in ['API', 'CAS', 'SMILES1', 'CCF', 'CAS', 'SMILES2', '预测标签说明']:
                    cell.alignment = left_align
                elif col_name in ['预测正类概率', '预测标签']:
                    cell.alignment = center_align
                    if col_name == '预测正类概率':
                        cell.number_format = '0.00%'
                else:
                    cell.alignment = right_align

            ws.row_dimensions[row].height = 22

        import unicodedata
        def display_width(text):
            return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))

        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_width = max((display_width(c.value) for c in col_cells if c.value is not None), default=0)
            final_width = max(8, min(max_width * 1.1 + 3, 50))
            ws.column_dimensions[letter].width = final_width

        ws.freeze_panes = 'A2'

        wb.save(OUTPUT_PATH)
        print(f"✅ 预测结果已保存至: {OUTPUT_PATH}")

    except Exception as e:
        print(f"❌ 结果保存失败: {e}")
        return

    print("\n" + "="*60)
    print("预测结果统计")
    print("="*60)
    print(f"总预测样本数: {len(result_df)}")
    print(f"预测可形成共晶的样本数: {sum(y_pred)}")
    print(f"预测无法形成共晶的样本数: {len(y_pred) - sum(y_pred)}")
    print(f"可形成共晶的占比: {sum(y_pred)/len(y_pred):.2%}")
    print("="*60)
    print("🎉 预测流程全部完成！")


if __name__ == "__main__":
    main()